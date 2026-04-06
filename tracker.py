"""
Application tracker. Maintains an xlsx spreadsheet logging every
resume generated, job details, scores, and status.
"""

import os
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


TRACKER_DIR = Path.home() / ".resume-mcp"
TRACKER_PATH = TRACKER_DIR / "applications.xlsx"

HEADERS = [
    "Date",
    "Company",
    "Job Title",
    "Job URL",
    "Status",
    "Overall Score",
    "Keyword Score",
    "Tech Score",
    "Impact Score",
    "Tech Matches",
    "Tech Gaps",
    "Keyword Gaps",
    "Resume ID",
    "PDF Path",
    "Notes",
]

COLUMN_WIDTHS = {
    "A": 12,   # Date
    "B": 20,   # Company
    "C": 25,   # Job Title
    "D": 40,   # Job URL
    "E": 14,   # Status
    "F": 14,   # Overall Score
    "G": 14,   # Keyword Score
    "H": 14,   # Tech Score
    "I": 14,   # Impact Score
    "J": 30,   # Tech Matches
    "K": 30,   # Tech Gaps
    "L": 30,   # Keyword Gaps
    "M": 12,   # Resume ID
    "N": 40,   # PDF Path
    "O": 30,   # Notes
}

# Styling
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SCORE_GOOD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
SCORE_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # yellow
SCORE_BAD = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # red

STATUS_COLORS = {
    "Generated": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "Optimized": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Applied": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "Interview": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Offer": PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_score_fill(score: int):
    if score >= 70:
        return SCORE_GOOD
    elif score >= 50:
        return SCORE_MED
    else:
        return SCORE_BAD


def _ensure_workbook() -> Workbook:
    """Load or create the tracker workbook."""
    TRACKER_DIR.mkdir(parents=True, exist_ok=True)

    if TRACKER_PATH.exists():
        return load_workbook(str(TRACKER_PATH))

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Set column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:O1"

    wb.save(str(TRACKER_PATH))
    return wb


def add_application(
    company: str,
    job_title: str,
    job_url: str = "",
    status: str = "Generated",
    overall_score: int = None,
    keyword_score: int = None,
    tech_score: int = None,
    impact_score: int = None,
    tech_matches: list[str] = None,
    tech_gaps: list[str] = None,
    keyword_gaps: list[str] = None,
    resume_id: int = None,
    pdf_path: str = None,
    notes: str = "",
) -> int:
    """Add a row to the tracker spreadsheet. Returns the row number."""
    wb = _ensure_workbook()
    ws = wb["Applications"]

    row = ws.max_row + 1

    data = [
        datetime.now().strftime("%Y-%m-%d"),
        company,
        job_title,
        job_url,
        status,
        overall_score,
        keyword_score,
        tech_score,
        impact_score,
        ", ".join(tech_matches or []),
        ", ".join(tech_gaps or []),
        ", ".join((keyword_gaps or [])[:10]),
        resume_id,
        pdf_path or "",
        notes,
    ]

    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Color-code scores
    for col_idx in [6, 7, 8, 9]:  # F, G, H, I
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.fill = _get_score_fill(cell.value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Color-code status
    status_cell = ws.cell(row=row, column=5)
    if status in STATUS_COLORS:
        status_cell.fill = STATUS_COLORS[status]

    # Make URL clickable
    url_cell = ws.cell(row=row, column=4)
    if job_url:
        url_cell.hyperlink = job_url
        url_cell.font = Font(color="0563C1", underline="single")

    wb.save(str(TRACKER_PATH))
    return row


def update_status(row: int, new_status: str, notes: str = None):
    """Update the status of an application by row number."""
    wb = _ensure_workbook()
    ws = wb["Applications"]

    status_cell = ws.cell(row=row, column=5)
    status_cell.value = new_status
    if new_status in STATUS_COLORS:
        status_cell.fill = STATUS_COLORS[new_status]

    if notes:
        ws.cell(row=row, column=15).value = notes

    wb.save(str(TRACKER_PATH))


def get_stats() -> dict:
    """Get summary statistics from the tracker."""
    wb = _ensure_workbook()
    ws = wb["Applications"]

    total = ws.max_row - 1  # exclude header
    if total <= 0:
        return {"total": 0}

    statuses = {}
    scores = []
    companies = set()

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=5).value
        score = ws.cell(row=row, column=6).value
        company = ws.cell(row=row, column=2).value

        if status:
            statuses[status] = statuses.get(status, 0) + 1
        if score:
            scores.append(score)
        if company:
            companies.add(company)

    return {
        "total_applications": total,
        "unique_companies": len(companies),
        "by_status": statuses,
        "avg_score": round(sum(scores) / len(scores), 1) if scores else None,
        "highest_score": max(scores) if scores else None,
        "tracker_path": str(TRACKER_PATH),
    }
